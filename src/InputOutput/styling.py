import pandas as pd


def _apply_xlsxwriter_header_styling(wb, ws) -> None:
    """Pas header styling toe voor xlsxwriter engine."""
    header_fmt = wb.add_format({"bold": True})
    ws.set_row(0, None, header_fmt)


def _apply_xlsxwriter_meta_column_styling(wb, ws, result_df: pd.DataFrame) -> None:
    """Pas meta kolom styling toe voor xlsxwriter engine."""
    grey_fmt = wb.add_format({"bg_color": "#DDDDDD"})
    
    # Meta kolommen styling
    meta_headers = ["Key"] + [c for c in result_df.columns if c.startswith("Aanwezig_")] + ["Match_Key"]
    header_to_index = {h: i for i, h in enumerate(result_df.columns)}
    
    for h in meta_headers:
        col_idx = header_to_index.get(h)
        if col_idx is not None:
            ws.set_column(col_idx, col_idx, None, grey_fmt)


def _apply_xlsxwriter_styling(wb, ws, result_df: pd.DataFrame) -> None:
    """Pas styling toe voor xlsxwriter engine."""
    _apply_xlsxwriter_header_styling(wb, ws)
    _apply_xlsxwriter_meta_column_styling(wb, ws, result_df)


def _apply_openpyxl_header_styling(ws) -> None:
    """Pas header styling toe voor openpyxl engine."""
    from openpyxl.styles import Font
    
    for cell in ws[1]:
        cell.font = Font(bold=True)


def _apply_openpyxl_meta_column_styling(ws, result_df: pd.DataFrame) -> None:
    """Pas meta kolom styling toe voor openpyxl engine."""
    from openpyxl.styles import PatternFill
    
    header_to_index = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}
    meta_headers = ["Key"] + [c for c in result_df.columns if c.startswith("Aanwezig_")] + ["Match_Key"]
    grey_fill = PatternFill(start_color="00DDDDDD", end_color="00DDDDDD", fill_type="solid")
    
    for h in meta_headers:
        col_idx = header_to_index.get(h)
        if col_idx is not None:
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    cell.fill = grey_fill


def _apply_openpyxl_styling(wb, ws, result_df: pd.DataFrame) -> None:
    """Pas styling toe voor openpyxl engine."""
    _apply_openpyxl_header_styling(ws)
    _apply_openpyxl_meta_column_styling(ws, result_df)


def apply_sheet_styling(writer: pd.ExcelWriter, sheet_name: str, result_df: pd.DataFrame) -> None:
    """Pas basisopmaak toe op headers en meta-kolommen."""
    if result_df.empty:
        return
    
    try:
        if writer.engine == "xlsxwriter":
            wb = writer.book
            ws = writer.sheets.get(sheet_name)
            if ws is not None:
                _apply_xlsxwriter_styling(wb, ws, result_df)
        else:
            wb = writer.book
            ws = wb[sheet_name]
            _apply_openpyxl_styling(wb, ws, result_df)
    except KeyError:
        # Sheet bestaat niet in workbook, sla styling over met milde waarschuwing
        import logging
        logging.getLogger("validation").warning("Styling overgeslagen: sheet '%s' niet gevonden", sheet_name)
    except Exception as exc:
        import logging
        logging.getLogger("validation").warning("Styling mislukt voor sheet '%s': %s", sheet_name, exc)


